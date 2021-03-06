from django.shortcuts import render, get_object_or_404
from catalog.models import Product, Supplier, Order, OrderItem, Basket, BasketResident, Resident, FedOrder, FedOrderItem, LimitFamily
from django.views import generic
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404
from django.urls import reverse, reverse_lazy
from django.db import transaction
from .forms import *
from dal import autocomplete
from django.core.mail import send_mail
from django.conf import settings
from django.contrib.auth.models import User, Group
from django.forms.utils import ErrorList
from catalog.mail import *
from django.core import serializers
from django.contrib import messages
import datetime
from django.contrib.auth.decorators import permission_required
from django.contrib.auth import authenticate, login
from django.db.models import Q, F
from django.utils import formats
import openpyxl
from pathlib import Path

LIST_OF_RESIDENT = []

def index(request):
	return render(request, 'index.html')

def ConsommationView(request):
    dict = {}
    op = "TOUS"
    if request.method == 'POST':
        op = request.POST.get('option')
        start_week= datetime.datetime.strptime(request.POST.get('startDate'), '%Y-%m-%d')
        end_week =datetime.datetime.strptime(request.POST.get('endDate'), '%Y-%m-%d')
        orderResident = Order.objects.filter(date__range=[start_week, end_week])
        orderFed = FedOrder.objects.filter(date__range=[start_week, end_week])
        orderitemResident = OrderItem.objects.filter(order__in=orderResident)
        orderitemFed = FedOrderItem.objects.filter(order__in=orderFed)
        if op == 'RESIDENT':
            prout = 'proutRESIDENT'
            for o in orderitemResident :
                if o.product.prod_name in dict :
                    value = dict.get(o.product.prod_name)
                    value += o.qty
                    dict[o.product.prod_name] = value
                else :
                    dict[o.product.prod_name] = o.qty
        elif  op == 'PERSONNEL':
            for o in orderitemFed :
                if o.product.prod_name in dict :
                    value = dict.get(o.product.prod_name)
                    value += o.qty
                    dict[o.product.prod_name] = value
                else :
                    dict[o.product.prod_name] = o.qty
        else : 
            for o in orderitemResident :
                if o.product.prod_name in dict :
                    value = dict.get(o.product.prod_name)
                    value += o.qty
                    dict[o.product.prod_name] = value
                else :
                    dict[o.product.prod_name] = o.qty
            for o in orderitemFed :
                if o.product.prod_name in dict :
                    value = dict.get(o.product.prod_name)
                    value += o.qty
                    dict[o.product.prod_name] = value
                else :
                    dict[o.product.prod_name] = o.qty
    else :  
        date = datetime.date.today()
        start_week = date - datetime.timedelta(date.weekday())
        end_week = start_week + datetime.timedelta(7)
        orderResident = Order.objects.filter(date__range=[start_week, end_week])
        orderFed = FedOrder.objects.filter(date__range=[start_week, end_week])
    return render(request, 'catalog/consommation.html', {'start_week':start_week,'end_week' : end_week,'op':op, 'dict':dict})
################################################################################################
#products
################################################################################################
def PrintOrder(request,pk):
    fedorder = get_object_or_404(FedOrder, pk=pk)
    return render(request,'catalog/printable_order.html', {'fedorder':fedorder,})
    
class ProductListView(generic.ListView):
    model = Product

def ProductListWarningView(request):
    product_list = Product.objects.filter(prod_stock__range=(1,F('prod_min')))
    return render(request,'catalog/product_list.html', {'product_list':product_list})

def ProductListOutView(request):
    product_list = Product.objects.filter(prod_stock='0')
    return render(request,'catalog/product_list.html', {'product_list':product_list})
    
def ProductListResidentView(request):
    user_basket = Resident.objects.all()
    products = Product.objects.filter(prod_main_category='RESIDENT')
    return render(request,'catalog/product_list_resident.html', {'product_list':products,'user_basket' : user_basket})
    
class ProductDetailView(LoginRequiredMixin, generic.DetailView):
    model = Product
    
def ProductListShop(request):
    products = Product.objects.filter(prod_main_category='PERSONNEL')
    return render(request,'catalog/product_list_shop.html', {'product_list':products})


def ProductCreate(request):
    if request.method == 'POST':
        form = ProductCreateForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('products') )
    else:
        form = ProductCreateForm()
    return render(request, 'catalog/product_create_form.html', {
            'form': form,
    })
    
def ProductUpdate(request, pk):
    instance = get_object_or_404(Product, pk=pk)
    form = ProductUpdateForm(request.POST or None, instance=instance)
    can_save = False
    product_list = Product.objects.exclude(pk=instance.pk)
    oldObject = get_object_or_404(Product, pk=pk)
    if form.is_valid():
        if oldObject.prod_ref_in != form.cleaned_data['prod_ref_in']:
            if product_list.filter(prod_ref_in=form.cleaned_data.get('prod_ref_in')).exists():
                form.add_error('prod_ref_in','La r??f??rence interne existe d??ja')
                return render(request, 'catalog/product_form.html', {'form': form}) 
        if oldObject.prod_ref_out != form.cleaned_data['prod_ref_out'] or oldObject.prod_supplier != form.cleaned_data['prod_supplier']:
            if product_list.filter(prod_supplier=form.cleaned_data.get('prod_supplier'),prod_ref_out=form.cleaned_data.get('prod_ref_out')).exists():
                form.add_error('prod_ref_out','La r??f??rence externe existe d??ja')
                return render(request, 'catalog/product_form.html', {'form': form}) 
        form.save()
        return HttpResponseRedirect(reverse('products') )
    return render(request, 'catalog/product_form.html', {'form': form}) 
    

class ProductDelete(DeleteView):
    model = Product
    success_url = reverse_lazy('products')

def ProductsBySupplierView(request, pk):
    model = Product
    products = Product.objects.filter(prod_supplier=pk)
    return render(request,'catalog/products_by_supplier.html', {'products': products})

class ProductAutocomplete(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        # Don't forget to filter out results depending on the visitor !
        if not self.request.user.is_authenticated:
           return Product.objects.none()

        qs = Product.objects.all()

        if self.q:
            qs = qs.filter(prod_name__istartswith=self.q)

        return qs 
        
def ProductRestock(request, pk):
    instance=get_object_or_404(Product, pk = pk)
    if request.method == 'POST':
        form = RestockProductForm(request.POST)
        if form.is_valid():
            instance.prod_stock += form.cleaned_data['qty_in']
            instance.save()
            return HttpResponseRedirect(reverse('products') )
    else:
        min = 1
        form = RestockProductForm(initial={'qty_in': min,})

    return render(request, 'catalog/restock_product_form.html', {'form': form, 'product':instance})
    
    
################################################################################################
#Basket
################################################################################################

def AddBasket(request,pk):
    if request.method == 'POST':
        error = []
        productselected = get_object_or_404(Product, pk=pk)
        if Basket.objects.filter(product=productselected, user_basket=request.user).exists():
            instancebasket = get_object_or_404(Basket, product=productselected, user_basket=request.user)
            qty = instancebasket.qty + int(request.POST.get('qty'))
            if qty <= productselected.prod_stock:
                instancebasket.qty += int(request.POST.get('qty'))
                instancebasket.save()
            else:
                qtymax = productselected.prod_stock - instancebasket.qty
                error.append("quantit?? maximum ajoutable : " + str(qtymax))
        else:
            instance = Basket(product=productselected , qty=request.POST.get('qty') , user_basket=request.user)
            instance.save()
        responseData = {
            'error': error,
        }
        return JsonResponse(responseData)
    else :
        return HttpResponseRedirect(reverse('product-shop') )
    
def AddBasketResident(request,pk):
    if request.method == 'POST':
        error = []
        productselected = get_object_or_404(Product, pk=pk)
        residentselected =  get_object_or_404(Resident, pk=request.POST.get('user_basket'))
        if BasketResident.objects.filter(product=productselected, user_basket=residentselected).exists():
            instancebasket = get_object_or_404(BasketResident, product=productselected, user_basket=residentselected)
            qty = instancebasket.qty + int(request.POST.get('qty'))
            if qty <= productselected.prod_stock:
                instancebasket.qty += int(request.POST.get('qty'))
                instancebasket.points = instancebasket.qty*productselected.prod_limit
                instancebasket.save()
            else:
                qtymax = productselected.prod_stock - instancebasket.qty
                error.append("quantit?? maximum ajoutable : " + str(qtymax))
        else:
            instance = BasketResident(product=productselected , qty=request.POST.get('qty') , user_basket=residentselected, points=int(request.POST.get('qty'))*productselected.prod_limit)
            instance.save()
        responseData = {
            'error': error,
        }
        return JsonResponse(responseData)
    else :
        return HttpResponseRedirect(reverse('product-shop') )
    
def BasketListView(request):
    product_list = Basket.objects.filter(user_basket=request.user)
    return render(request,'catalog/basket.html', {'product_list':product_list})
    
def BasketResidentListView(request,pk):
    product_list = BasketResident.objects.filter(user_basket=pk)
    points = 0
    for p in product_list :
        points += p.product.prod_limit * p.qty
    instance = get_object_or_404(Resident, pk=pk)
    date = datetime.date.today()
    start_week = date - datetime.timedelta(date.weekday())
    end_week = start_week + datetime.timedelta(7)
    family = Resident.objects.filter(badge=instance.badge)
    pklist = []
    for f in family:
        pklist.append(f.id)
    entries = Order.objects.filter(date__range=[start_week, end_week], order_user__in=pklist)
    pointrestant =0
    
    for p in entries :
        pointrestant += p.points
    family = Resident.objects.filter(badge=instance.badge)
    familynumber = family.count()
    user_points = get_object_or_404(LimitFamily, compo_family=str(familynumber)).point_by_week
    pointrestant = user_points - pointrestant
    ptwithorder = pointrestant - points
    
    return render(request,'catalog/basket_resident_detail.html', {'product_list':product_list, 'points':points, 'user_points':user_points, 'resident':instance,'pointrestant':pointrestant,'ptwithorder':ptwithorder})
    
def BasketDelete(request):
    product_list = Basket.objects.filter(user_basket=request.user)
    for p in product_list :
        p.delete()
    return HttpResponseRedirect(reverse('basket') )

def BasketResidentDelete(request,pk):
    instance = get_object_or_404(Resident, pk=pk)
    product_list = BasketResident.objects.filter(user_basket=instance)
    for p in product_list :
        p.delete()
    return HttpResponseRedirect(reverse('orders') )
    
def BasketConvert(request):
    product = Product.objects.all()
    order_list = Basket.objects.filter(user_basket=request.user)
    can_save = True
    error = []
    for p in order_list :
        product = get_object_or_404(Product, pk=p.product.pk)
        if p.qty > product.prod_stock:
            error.append("Stock insuffisant - Stock restant :"+ str(product.prod_stock))
            p.save()
            can_save = False
    if not Basket.objects.filter(user_basket=request.user).exists() : 
        error.append("panier vide")
        can_save = False
        
    if can_save:
        id = FedOrder.objects.latest('id')
        title = "AMC_Fed_"+datetime.datetime.now().strftime ("%d%m%Y")+"_"+str((id.id+1))
        instance_order = FedOrder(order_user=request.user,title=title,commentary=request.POST.get('commentary'))
        instance_order.save()
        for p in order_list :
            instance_item = FedOrderItem(product=p.product,order=instance_order,qty=p.qty, qty_supplied=p.qty)
            instance_item.save()
            p.delete()
        email_new_order(instance_order)
    responseData = {
        'error': error,
    }
    return JsonResponse(responseData)
    #return HttpResponseRedirect(reverse('basket') )
 
def BasketResidentConvert(request, pk):
    product = Product.objects.all()
    instance = get_object_or_404(Resident, pk=pk)
    order_list = BasketResident.objects.filter(user_basket=instance)
    verifpoint = request.POST.get('verifpoint')
    date = datetime.date.today()
    start_week = date - datetime.timedelta(date.weekday())
    end_week = start_week + datetime.timedelta(7)
    family = Resident.objects.filter(badge=instance.badge)
    pklist = []
    for f in family:
        pklist.append(f.id)
    entries = Order.objects.filter(date__range=[start_week, end_week], order_user__in=pklist)
    point =0
    
    for p in entries :
        point += p.points
    pts=0
    can_save = True
    error = []
    if not BasketResident.objects.filter(user_basket=instance).exists() : 
        error.append("panier vide")
        can_save = False
    else :    
        for p in order_list :
            pts += p.product.prod_limit * p.qty
            if p.qty > p.product.prod_stock:
                error.append("Stock insuffisant de : "+p.product.prod_name +" - Stock restant :"+ str(p.product.prod_stock))
                p.save()
                can_save = False
                
        if verifpoint:       
            instance = get_object_or_404(Resident, pk=pk)
            family = Resident.objects.filter(badge=instance.badge)
            familynumber = family.count()
            user_points = get_object_or_404(LimitFamily, compo_family=str(familynumber)).point_by_week
            restant = user_points - point
            if pts > restant:
                error.append("limite de point d??pass??, point restant :" + str(restant))
                can_save = False
        else:
            pts=0
            
       
    
    if can_save:
       
        #for p in order_list :
            
        id = Order.objects.latest('id')
        title = "AMC_Shop_"+datetime.datetime.now().strftime ("%d%m%Y")+"_"+str((id.id+1))
        instance_order = Order(order_user=instance,title=title,points = pts)
        instance_order.save()
        for p in order_list :
            instance_item = OrderItem(product=p.product,order=instance_order,qty=p.qty)
            instance_item.save()
            p.delete()
    responseData = {
        'error': error,
        'verifpoint': verifpoint,
        #'entries':list(entries.values()),
        'point':point,
    }
    return JsonResponse(responseData)
    # return HttpResponseRedirect(reverse('orders') )

def ProductAddOne(request, pk):
    instance=get_object_or_404(Basket, pk = pk)
    error = []
    if request.method == 'POST':
        if instance.qty+1 <= instance.product.prod_stock :
            instance.qty += 1
            instance.save()
        else:
            error.append('La quantit?? de : '+ instance.product.prod_name +' est ??gale ?? la quantit?? en stock')
    responseData = {
        'error': error,
    }
    return JsonResponse(responseData)            

def BasketResidentAddOne(request, pk):
    instance=get_object_or_404(BasketResident, pk = pk)
    error = []
    if request.method == 'POST':
        if instance.qty+1 <= instance.product.prod_stock :
            instance.qty += 1
            instance.points = instance.qty*instance.product.prod_limit
            instance.save()
        else :
            error.append('La quantit?? de : '+ instance.product.prod_name +' est ??gale ?? la quantit?? en stock')
    responseData = {
        'error': error,
    }
    return JsonResponse(responseData)
    
def BasketResidentRemoveOne(request, pk):
    instance=get_object_or_404(BasketResident, pk = pk)

    if request.method == 'POST':
        if instance.qty - 1 > 0 :
            instance.qty -= 1
            instance.points = instance.qty*instance.product.prod_limit
            instance.save()
        elif instance.qty - 1 == 0:
            instance.delete()
    return HttpResponseRedirect(reverse('orders') )
    
def ProductRemoveOne(request, pk):
    instance=get_object_or_404(Basket, pk = pk)

    if request.method == 'POST':
        if instance.qty - 1 > 0 :
            instance.qty -= 1
            instance.save()
        elif instance.qty - 1 == 0:
            instance.delete()
    return HttpResponseRedirect(reverse('basket') )
    
################################################################################################
#Commande Personnel
################################################################################################

def FedOrderListView(request):
    fedorder_list_open = FedOrder.objects.filter(Q(status='PARTIEL') | Q(status='OPEN'))
    fedorder_list_close = FedOrder.objects.filter(status='CLOSED')
    return render(request, 'catalog/fedorder_list.html', {'fedorder_list': fedorder_list_open, 'fedorder_list_closed':fedorder_list_close})

class FedOrderDelete(DeleteView):
    model = FedOrder
    success_url = reverse_lazy('fed-orders')
    
def MyFedOrderListView(request):
    fedorder_list = FedOrder.objects.filter(Q(order_user=request.user, status='PARTIEL') | Q(order_user=request.user, status='OPEN'))
    fedorder_list_close = FedOrder.objects.filter(order_user=request.user, status='CLOSED')
    return render(request, 'catalog/fedorder_list.html', {
            'fedorder_list': fedorder_list,
            'fedorder_list_closed':fedorder_list_close
    })

class FedOrderDetailView(LoginRequiredMixin, generic.DetailView):
    model = FedOrder

@permission_required('catalog.can_close_order')
def FedOrderClose(request, pk):
    instance = get_object_or_404(FedOrder, pk=pk)
    fedorderitem = FedOrderItem.objects.filter(order = instance)
    nbitem = FedOrderItem.objects.filter(order = instance).count()
    count = 0
    orderin = 0
    error = []
    can_save = True
    for f in fedorderitem:
        if f.delivered and not f.already_delivered :
            instance_product = get_object_or_404(Product, pk = f.product.pk)
            if instance_product.prod_stock < f.qty_supplied :
                error.append('Le produit : "'+instance_product.prod_name+"\" n'a plus que : "+ str(instance_product.prod_stock) +" en stock")
                can_save = False
    if can_save:
        for f in fedorderitem:
            orderin += 1
            if f.already_delivered :
                count += 1
            if f.delivered and not f.already_delivered :
                instance_product = get_object_or_404(Product, pk = f.product.pk)
                if instance_product.prod_stock > f.qty_supplied :
                    instance_product.prod_stock = instance_product.prod_stock-f.qty_supplied
                    instance_product.save()
                    f.already_delivered = True
                    f.save()
                    count += 1
    if count != 0 :
        if count == orderin :
            instance.status = "CLOSED"
        else : 
            instance.status = "PARTIEL"
        instance.prepared_by = request.user
        instance.save()
    fedorder_list = FedOrder.objects.all()
    responseData = {
        'error': error,
    }
    return JsonResponse(responseData)

# @permission_required('catalog.can_close_order')
# def FedOrderPartiel(request, pk):
    # instance = get_object_or_404(FedOrder, pk=pk)
    # instance.status = "PARTIEL"
    # instance.prepared_by = request.user
    # instance.save()
    # fedorderitem = FedOrderItem.objects.filter(order = instance)
    # for f in fedorderitem:
        # instance_product = get_object_or_404(Product, pk = f.product.pk)
        # instance_product.prod_stock = instance_product.prod_stock-f.qty_supplied
        # instance_product.save()
    # fedorder_list = FedOrder.objects.all()
    # return HttpResponseRedirect(reverse('fed-orders'))

def FedOrderUpdateView(request, pk):
    instance = get_object_or_404(FedOrder, pk=pk)
    return render(request, 'catalog/fedorder_update.html', {
            'order': instance,
    })
    
def FedOrderItemUpdateView(request,pk):
    instance = get_object_or_404(FedOrderItem, pk=pk)
    instance.qty_supplied = request.POST.get('qty')
    if request.POST.get('unchecked') == '1' :
        instance.delivered = True
    else : 
        instance.delivered = False
    instance.save()
    return render(request, 'catalog/fedorder_update.html', {
            'fedorderitem': instance,
    })
   
################################################################################################
#suppliers
################################################################################################

class SupplierListView(LoginRequiredMixin, generic.ListView):
    model = Supplier
     
class SupplierDetailView(LoginRequiredMixin, generic.DetailView):
    model = Supplier

def SupplierCreate(request):
    if request.method == 'POST':
        form = SupplierCreateForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('suppliers') )
    else:
        form = SupplierCreateForm()
    return render(request, 'catalog/supplier_create_form.html', {
            'form': form,
    })

def SupplierUpdate(request, pk):
    obj= get_object_or_404(Supplier, id=pk)
        
    form = SupplierCreateForm(request.POST or None, instance= obj)
    context= {'form': form}

    if form.is_valid():
        obj= form.save(commit= False)
        obj.save()
        messages.success(request, "You successfully updated the post")
        context= {'form': form}
        return HttpResponseRedirect(reverse('suppliers') )
    else:
        context= {'form': form,
                           'error': 'The form was not updated successfully. Please enter in a title and content'}
        # return HttpResponseRedirect(reverse('suppliers') )
        return render(request, 'catalog/supplier_update_form.html', {'form': form}) 

class SupplierDelete(DeleteView):
    model = Supplier
    success_url = reverse_lazy('suppliers')
    
################################################################################################    
#residents
################################################################################################   

class ResidentListView(LoginRequiredMixin, generic.ListView):
    model = Resident
    
class ResidentDetailView(LoginRequiredMixin, generic.DetailView):
    model = Resident

def ResidentDetailFamilyView(request, pk):
    model = Resident
    instance = get_object_or_404(Resident, pk=pk)
    family = Resident.objects.filter(badge=instance.badge).order_by('-age')
    familynumber = family.count()
    child = 0
    for f in family:
        if f.age < 12:
            child += 1
    return render(request,'catalog/family_detail.html', {'child':child, 'resident': instance, 'family' : family, 'familynumber':familynumber})
    
def ResidentDetailOrderView(request, pk):
    model = Resident
    instance = get_object_or_404(Resident, pk=pk)
    family = Resident.objects.filter(badge=instance.badge)
    pklist = []
    for f in family:
        pklist.append(f.id)
    order_list = Order.objects.filter(order_user__in=pklist)
    return render(request,'catalog/family_detail_order.html', {'order_list':order_list})

class ResidentDelete(DeleteView):
    model = Resident
    success_url = reverse_lazy('residents')
    
def ResidentCreate(request):
    if request.method == 'POST':
        form = ResidentCreateForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('resident') )
    else:
        form = ResidentCreateForm()
    return render(request, 'catalog/resident_create_form.html', {
            'form': form,
    })
################################################################################################    
#Orders
################################################################################################   
    

    
class OrderListView(LoginRequiredMixin, generic.ListView):
    model = Order
    
class OrderDetailView(LoginRequiredMixin, generic.DetailView):
    model = Order

class OrderDelete(DeleteView):
    model = Order
    success_url = reverse_lazy('orders')
    
class OrderItemCreate(CreateView):
    form_class = OrderForm
    template_name = 'catalog/order_form.html'
    # model = Order
    # fields = '__all__'
    success_url = reverse_lazy('orders')
    
    def get_context_data(self, **kwargs):
        data = super(OrderItemCreate, self).get_context_data(**kwargs)
        if self.request.POST:
            data['orderlist'] = OrderItemFormSet(self.request.POST)
            
        else:
            data['orderlist'] = OrderItemFormSet()
        return data
    
    def form_valid(self, form):
        context = self.get_context_data(form=form)
        orderlist = context['orderlist']
        # residentselected = context['form'].cleaned_data.get('order_user')      
        # resident = get_object_or_404(Resident, pk=residentselected.pk)
        # family = Resident.objects.filter(badge=resident.badge).order_by('-age')
        # familynumber = family.count()
        # limit = get_object_or_404(LimitFamily, compo_family=str(familynumber))
        can_save = True
        d = dict()
        point = 0
        for f in orderlist:
            if f.is_valid():
                if f.cleaned_data.get('product') is None :
                    form.add_error(None,'Aucun produit s??lectionner')
                    can_save = False
                if f.cleaned_data.get('qty') is None :
                    form.add_error(None,'La quantit?? ne peut pas ??tre nulle')
                    can_save = False
                else : 
                    prod = f.cleaned_data['product']
                    point += prod.prod_limit
                    if prod.id in d :
                        qty = d.get(prod.id)
                        qty += f.cleaned_data['qty']
                        d[prod.id] = qty
                    else:
                        qty = f.cleaned_data['qty']
                        d[prod.id] = qty

        form.data = form.data.copy()
        form.data['points'] = 50
        orderlist.data = orderlist.data.copy()
        orderlist.data['points'] = 50
        # if point > limit.point_by_week : 
            # form.add_error(None,'Limite de point d??passe de : '+str(point))
            # can_save = False
            # return super().form_invalid(form)
        if can_save and orderlist.is_valid():
            
            response = super().form_valid(form)
            orderlist.instance = self.object
            orderlist.save()
            for key in d :
                product = Product.objects.get(pk=key)
                product.prod_stock = product.prod_stock-d.get(key)
                product.save()
                # if product.prod_stock <= product.prod_min:
                    # email_warning_stock(product)
            return response
        else:
            return super().form_invalid(form)
        return super(OrderItemCreate, self).form_valid(form)
 
        
class OrderUpdate(UpdateView):
    model = Order
    fields = '__all__'

    def get_context_data(self, **kwargs):
        data = super(OrderUpdate, self).get_context_data(**kwargs)
        if self.request.POST:
            data['orderlist'] = OrderItemFormSet(self.request.POST, instance=self.object)
        else:
            data['orderlist'] = OrderItemFormSet(instance=self.object)
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        orderlist = context['orderlist']
        if orderlist.is_valid():
            response = super().form_valid(form)
            orderlist.instance = self.object
            orderlist.save()
            return response
        else:
            return super().form_invalid(form)
        return super(OrderUpdate, self).form_valid(form)

    def get_success_url(self):
        return reverse_lazy('order-detail', kwargs={'pk': self.object.pk})
        
def OrderLastID(request):
    model = Order
    instance = Order.objects.latest('id')
    return HttpResponse(instance.id+1)


    
def Import(request) :

    responseData ={}
    select = ""
    dict =[]
    nbdays = 0
    ListOfResidentJSON =[]
    # LIST_OF_RESIDENT = []
    ListOfResident =[]
    if request.method == 'POST':
        delete = []
        select = request.POST.get('AddOrRemove')
        if select == "Add" :
            wb_obj = openpyxl.load_workbook(request.FILES['filexlsx'])
            sheet = wb_obj.active
            for row in sheet.iter_rows(max_col=9):
                value = []
                for cell in row:
                    if cell.value is None : 
                        break
                    elif cell.value == "Badge" or Resident.objects.filter(badge=str(cell.value)+".0").exists():
                        break
                    else:
                        value.append(cell.value)
                if value:
                    dict.append(value)
            for i in dict :
                resident = Resident(badge=str(i[0])+".0",family_group=i[1],name=i[2],firstname=i[3],room=i[4],age=i[6],datein=i[7],sexe=i[8])
                # today = datetime.datetime.now()
                # orig_date = str(resident.datein)
                # d2 = datetime.datetime.strptime(orig_date, "%Y-%m-%d %H:%M:%S")
                # d2 = d2.strftime('%Y-%m-%d')
                # d2 = datetime.datetime.strptime(d2, '%Y-%m-%d')
                # today = str(today)
                # d1 = datetime.datetime.strptime(today, "%Y-%m-%d %H:%M:%S.%f")
                # d1 = d1.strftime('%Y-%m-%d')
                # d1 = datetime.datetime.strptime(d1, '%Y-%m-%d')
                # nbdays = d1 - d2
                serialized_obj = serializers.serialize('json', [resident,])
                ListOfResidentJSON.append(serialized_obj)
                ListOfResident.append(resident)
                # LIST_OF_RESIDENT = ListOfResident
            request.session['import'] = serializers.serialize('json', ListOfResident)
            # responseData = {
                # 'select':select,
                # 'ListOfResident':ListOfResident,
                # 'nbdays':nbdays.days
             # }
            # for i in line:
                # delete.append(i+'.0')
            # Resident.objects.filter(badge__in=delete).delete()
            return render(request, 'catalog/resident_list.html', {
                'resident_list': ListOfResident, 'action':select
            })
        else :
            wb_obj = openpyxl.load_workbook(request.FILES['filexlsx'])
            sheet = wb_obj.active
            for row in sheet.iter_rows(max_col=2):
                value = ""
                for cell in row:
                    if cell.value is None : 
                        break
                    elif cell.value == "Badge" or cell.value in dict:
                        break
                    elif Resident.objects.filter(badge=str(cell.value)+".0").exists():
                        value = str(cell.value)+".0"
                if value:
                    dict.append(value)
            # print(dict)
            ListOfResident = Resident.objects.filter(badge__in=dict)   
            request.session['import'] = serializers.serialize('json', ListOfResident)            
            # responseData = {
                # 'select':select,
                # 'ListOfResident':request.session['import'],
                # # 'nbdays':nbdays.days
             # }
        return render(request, 'catalog/resident_list.html', {
                'resident_list': ListOfResident, 'action':select
            })
        
    else :
        return render(request, 'catalog/import.html')

def AddImport(request):
    for obj in serializers.deserialize('json',request.session['import']):
        obj.object.save()
    list = request.session['import']
    del request.session['import']
    return HttpResponseRedirect(reverse('residents') )


def DeleteImport(request):
    for obj in serializers.deserialize('json',request.session['import']):
        obj.object.delete()
    list = request.session['import']
    del request.session['import']
    return HttpResponseRedirect(reverse('residents') )

def DateIn(request,pk):
    resident = get_object_or_404(Resident, pk=pk)
    if resident.datein is not None:
        today = datetime.datetime.now()
        orig_date = str(resident.datein)[:19]
        d2 = datetime.datetime.strptime(orig_date, "%Y-%m-%d %H:%M:%S")
        d2 = d2.strftime('%Y-%m-%d')
        d2 = datetime.datetime.strptime(d2, '%Y-%m-%d')
        today = str(today)
        d1 = datetime.datetime.strptime(today, "%Y-%m-%d %H:%M:%S.%f")
        d1 = d1.strftime('%Y-%m-%d')
        d1 = datetime.datetime.strptime(d1, '%Y-%m-%d')
        days = d1 - d2
        nbdays = days.days
    else : 
        nbdays = "Pas de date d'arriv??e"
    responseData = {
        'nbdays': nbdays,
    }
    return JsonResponse(responseData)
        
class ResidentAutocomplete(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        if not self.request.user.is_authenticated:
           return Resident.objects.none()
        qs = Resident.objects.all()
        if self.q:
            qs = qs.filter(badge__istartswith=self.q)
        return qs

    
def register_request(request):
    form = NewUserForm()
    if request.method == "POST":
        form = NewUserForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, "Registration successful." )
            return HttpResponseRedirect(reverse('products'))
        form.add_error(None, "Unsuccessful registration. Invalid information.")
    return render (request=request, template_name="register.html", context={"form":form})
   
# def email_new_order(request):
    # users = User.objects.filter(groups__name='Shop Members')
    # for user in users:
        # subject = 'Nouvelle commande'
        # message = ' Une nouvelle commande est arriv??e'
        # email_from = settings.EMAIL_HOST_USER
        # recipient_list = [user.email,]
        # send_mail( subject, message, email_from, recipient_list )